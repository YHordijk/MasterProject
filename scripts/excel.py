import openpyxl as pxl
from win32com.client import Dispatch
import paths, os


#colors
fillRed = 		pxl.styles.PatternFill(start_color='FFFFD3D3', end_color='FFFFD3D3', fill_type='solid')
fillPurple = 	pxl.styles.PatternFill(start_color='FFE9C5FF', end_color='FFE9C5FF', fill_type='solid')
fillGreen = 	pxl.styles.PatternFill(start_color='FFE2EFDA', end_color='FFE2EFDA', fill_type='solid')
fillYellow = 	pxl.styles.PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid')
fillGrey = 		pxl.styles.PatternFill(start_color='FFD6DCE4', end_color='FFD6DCE4', fill_type='solid')
status_fills = {'Success': fillGreen, 'Failed':fillRed, 'Running':fillPurple, 'Warning':fillYellow, 'Queued':fillGrey, 'Canceled':fillRed}
text_colors = {'Success': 'FF375623', 'Failed':'FFC00000', 'Running':'FF7030A2', 'Warning':'FF80600B', 'Queued':'FF222B35', 'Canceled':'FFC00000'}

#border
bSides = pxl.styles.borders.Border(left=pxl.styles.borders.Side(style='thin'), right=pxl.styles.borders.Side(style='thin'))
bHeader = pxl.styles.borders.Border(left=pxl.styles.borders.Side(style='thin'), right=pxl.styles.borders.Side(style='thin'), bottom=pxl.styles.borders.Side(style='double'))

wb = pxl.Workbook()
wb.remove(wb.active) #remove default sheet
wb.create_sheet(title='Overview')
ws = wb['Overview']
with open(paths.results_table) as csv:
	lines = csv.readlines()
	fieldnames = [n.strip() for n in lines[1].split(',')]
	data = [[x.strip() for x in d.split(',')] for d in lines[2:]]

	for i, n in enumerate(fieldnames):
		row = 1
		col = i+1
		ws.cell(column=col, row=row, value=n)
		ws.cell(column=col, row=row).font = pxl.styles.Font(bold=True)
		ws.cell(column=col, row=row).border = bHeader
		if n == 'DIRECTORY':
			directory_idx = i
		if n == 'STATUS':
			status_idx = i


	directories = [d[directory_idx] for d in data]
	statuses = [d[status_idx] for d in data]
	Nrows = len(data)
	Ncols = len(data[0])
	for i, d in zip(range(Nrows), data):
		row = i + 2
		status_fill = status_fills[statuses[i]]
		status_text_col = text_colors[statuses[i]]
		for j, n, x in zip(range(Ncols), fieldnames, d):
			col = j+1
			cell = ws.cell(column=col, row=row)
			if j == 0:
				cell.font = pxl.styles.Font(bold=True)
			if n == 'DIRECTORY':
				cell.value = f'=HYPERLINK("{os.path.join(paths.results, x)}", "{x}")'
				cell.style = 'Hyperlink'
			elif n == 'OUTXYZ':
				if not x == '':
					cell.value = f'=HYPERLINK("{os.path.join(paths.master,directories[i])}/molview2_start_out.bat", "output.xyz")'
					cell.style = 'Hyperlink'
			elif n == 'INXYZ':
				if not x == '':
					cell.value = f'=HYPERLINK("{os.path.join(paths.master,directories[i])}/molview2_start_in.bat", "input.xyz")'
					cell.style = 'Hyperlink'
			elif n == 'RUNTIME':
				cell.value = x
				cell.alignment = pxl.styles.Alignment(horizontal='right')
			else:
				cell.value = x

			if n in ['ID', 'STATUS']:
				cell.fill = status_fill
				cell.font = pxl.styles.Font(color=status_text_col, bold=True)
			cell.border = bSides



last_col = pxl.utils.cell.get_column_letter(col)
ws.auto_filter.ref = f"A1:{last_col}{row}"
ws.freeze_panes = ws['B2']


wb.save(paths.results_table_pretty)

excel = Dispatch('Excel.Application')
wb = excel.Workbooks.Open(paths.results_table_pretty)
excel.Worksheets(1).Activate()
excel.ActiveSheet.Columns.AutoFit()
wb.Save()
wb.Close()